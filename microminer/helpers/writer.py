import pandas as pd
import os
from openpyxl import load_workbook

def write_metrics_to_excel(data, file_path, sheet_name='Metrics'):
    # Convert the data into a DataFrame
    df = pd.DataFrame([data])

    # Check if file exists
    if os.path.exists(file_path):
        # Load the workbook and the sheet
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        # Check if the sheet exists
        if sheet_name in book.sheetnames:
            # Load the existing sheet
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            startrow = writer.sheets[sheet_name].max_row
            
            # Append data without overwriting the existing content
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
        else:
            # Add new sheet with data
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Save the file
        writer.save()
    else:
        # Write new file if it doesn't exist
        df.to_excel(file_path, sheet_name=sheet_name, index=False)